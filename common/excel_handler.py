"""读取用例模块"""
import os

import openpyxl
from collections import namedtuple

from config import path


class CaseData(object):
    def __init__(self, zip_obj, *args, **kwargs):
        """测试用例数据类"""
        for i in list(zip_obj):
            setattr(self, i[0], i[1])


# 定义一个类, 专门用来读取excel中的数据
class ReadExcel(object):
    """读取excel中的用例数据"""

    def __init__(self, file_name, sheet_name):
        """
        :param file_name: excel文件名
        :param sheet_name: sheet表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开工作簿和表单"""
        # 打开文件, 返回一个工作簿对象
        self.workbook = openpyxl.load_workbook(self.file_name)
        # 通过工作簿, 选择表单对象
        self.sheet = self.workbook[self.sheet_name]

    def close(self):
        # 关闭工作簿
        self.workbook.close()

    def read_data_obj(self):
        # 打开工作簿
        self.open()
        # 创建空的列表,用来存放所有的的用例数据
        cases = []
        # 读取表单中的数据
        rows = list(self.sheet.rows)
        # 读取表头
        titles = []
        for i in rows[0]:
            titles.append(i.value)

        for row in rows[1:]:
            # 读每一行数据
            case = []
            for r in row:
                if  r.value != None:
                    case.append(r.value)
            zip_obj = zip(titles, case)
            # 将每一条用例的数据, 存储为一个对象
            # 通过Case这个类来创建一个对象, 传了一个参数, zip_obj
            case_data = CaseData(zip_obj)
            cases.append(case_data)

        # 将包含所偶用例的列表cases进行返回
        return cases

    def write_data(self, row, column, value):
        """
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的内容
        :return:
        """
        # 打开文件
        self.open()
        # 按照传入的行, 列, 内容进行写入
        self.sheet.cell(row=row, column=column, value=value)
        # 保存
        self.workbook.save(self.file_name)
        self.close()

    def get_max_row(self):
        '''表单最大行不准时 使用这个方法获取最大行'''
        self.open()
        i = self.sheet.max_row
        real_max_row = 0
        while i > 0:
            print(self.sheet[i])
            row_dict = {i.value for i in self.sheet[i]}
            if row_dict == None:
                i = i - 1
            else:
                real_max_row = i
                break
        return real_max_row

    def read_data_tuple(self):
        self.open()
        """
        按行读取excel中的数据，以列表的形式返回，在生成一个命名元祖对象与之绑定   提高性能
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for index, title in enumerate(rows_data[0]):  # 将非空的列表头取出
            if rows_data[1][index].value:
                titles.append(title.value)
        # 定义一个空列表用来存储所有的命名元组
        cases_list = []
        for case in rows_data[1:]:
            #     # data用例临时存放用例数据
            data = []
            for cell in case:  # 遍历所有表体
                if isinstance(cell.value, str):  # 转换添加
                    if cell.value != None:
                        data.append(cell.value)
                else:
                    if cell.value != None:
                        data.append(str(cell.value))

            if len(data) == 0:  # 如果遍历出来没有需要的数据跳出循环
                break
            cases = namedtuple("cases", titles)
            cases = cases(*data)
            cases_list.append(cases)
        return cases_list

    def readline_data_obj(self, list1=None):
        """
        按指定的列，读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
        excel中的表头为对象的属性，对应的数据为属性值
        :param list1: list  -->要读取的列[1,2....]
        :return: type:list--->[case_obj1,case_obj2....]，
        """
        self.open()
        list1 = eval(list1)
        if len(list1) == 0:
            return self.read_data_obj()

        max_r = self.sheet.max_row
        # 定义一个空列表，用来存放所有用例
        cases = []
        # 定义一个空列表，用来存放表头
        titles = []
        # 遍历出所有的行
        for row in range(1, max_r + 1):
            # 判断是否是第一行
            if row == 1:
                # 遍历list1指定的列，获取表头
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    # 将数据添加到表头中
                    titles.append(title)
            else:
                # 定义一个空列表，用来存放该行的数据
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                # 将该条数据和表头进行打包组合，
                case = dict(zip(titles, case_data))
                # 创建一个用例对象，将表头和数据传进入初始化

                cases.append(case)
        self.close()
        return cases


    def read_data_obj2(self):
        # 打开工作簿
        self.open()
        # 创建空的列表,用来存放所有的的用例数据
        cases = []
        # 读取表单中的数据
        rows = list(self.sheet.rows)
        # 读取表头
        titles = []
        for i in rows[0]:
            titles.append(i.value)
        for row in rows[1:]:
            # 读每一行数据
            case = []
            for r in row:
                # if  r.value != None:
                    case.append(r.value)
            zip_obj = dict(zip(titles, case))
            # 将每一条用例的数据, 存储为一个对象
            # 通过Case这个类来创建一个对象, 传了一个参数, zip_obj
            # case_data = CaseData(zip_obj)
            cases.append(zip_obj)

        # 将包含所偶用例的列表cases进行返回
        self.close()
        return cases







